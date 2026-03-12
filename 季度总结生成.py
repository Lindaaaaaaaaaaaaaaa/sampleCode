

import pandas as pd
from datetime import timedelta
import xlwings as xw
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
import os
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from matplotlib.dates import MonthLocator, DateFormatter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl import load_workbook
import re
import os

def merge_second_sheets(input_folder, output_file):
    """
    合并指定文件夹内所有Excel文件的第二张工作表
    :param input_folder: 包含Excel文件的文件夹路径
    :param output_file: 合并结果输出路径
    """
    merged_data = pd.DataFrame()
    
    for filename in os.listdir(input_folder):
        if not filename.lower().endswith(('.xlsx', '.xls')):
            continue
            
        try:
            filepath = os.path.join(input_folder, filename)
            wb = load_workbook(filepath, read_only=True)
            sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else None
            
            if sheet_name:
                print(f"正在合并：{sheet_name}")
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                # 在合并前先删除指定列
                df = df.drop(['买入交易量（亿元）', '卖出交易量（亿元）'], axis=1, errors='ignore')
                merged_data = pd.concat([merged_data, df], ignore_index=True)
                
        except Exception as e:
            print(f"处理文件 {filename} 时出错: {str(e)}")
    
    merged_data.to_excel(output_file, index=False)
    print(f"合并完成！结果已保存至 {output_file}")

    
def set_excel_dimensions(excel_path, row_idx, img_height, columns, img_width):
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 批量设置行高（1到row_index行）
    for i in range(1, row_idx + 1):
        ws.row_dimensions[i].height = img_height * 0.75  # 行高单位转换
    
    # 批量设置列宽
    for col_letter in columns:
        ws.column_dimensions[col_letter].width = max(
            img_width / 7,  # 基础宽度计算
            ws.column_dimensions[col_letter].width  # 保留原值防缩小
        )
    
    wb.save(excel_path)


def insert_images_to_excel(excel_path, image_folder, target_cell="A1", img_width=100, img_height=100, num=7):
    """改进版Excel图片插入程序，支持多列和自动换列"""
    try:
        # 参数校验
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
        if not os.path.isdir(image_folder):
            raise NotADirectoryError(f"图片文件夹无效: {image_folder}")

        # 加载工作簿
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # 解析初始单元格坐标
        col_letter = re.sub(r'[^A-Za-z]', '', target_cell)
        row_num = int(re.sub(r'[^0-9]', '', target_cell))
        start_row = row_num
        
        # 获取图片文件列表
        image_files = [f for f in sorted(os.listdir(image_folder)) 
                      if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        
        if not image_files:
            print("警告: 未找到任何图片文件")
            return

        # 插入图片
        for i, filename in enumerate(image_files):
            try:
                img_path = os.path.join(image_folder, filename)
                img = ExcelImage(img_path)
                img.width, img.height = img_width, img_height
                
                # 计算当前插入位置
                current_col = chr(ord(col_letter) + i // num)
                current_row = start_row + i % num
                
                ws.add_image(img, f"{current_col}{current_row}")
                print(f"成功插入: {filename} -> {current_col}{current_row}")
                
            except Exception as e:
                print(f"插入失败 {filename}: {str(e)}")
                continue
        
        # 保存文件
        base, ext = os.path.splitext(excel_path)
        output_path = f"{base}_with_images{ext}"
        wb.save(output_path)
        print(f"\n处理完成! 结果已保存至: {output_path}")
        
    except Exception as e:
        print(f"\n程序出错: {str(e)}")


def daily_to_weekly(input_file, output_file):
    # 数据读取与预处理
    print("数据提取中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    
    if '净买入交易量（亿元）' in df.columns:
        df['净买入交易量（亿元）'] = safe_convert_to_numeric(df['净买入交易量（亿元）'])
    
    df['交易日期'] = pd.to_datetime(df['交易日期'], format='%Y/%m/%d', errors='coerce')
    df = df[df['交易日期'].dt.weekday < 5]  # 关键修改1：过滤周末
    
    df['周分组'] = df['交易日期'].apply(
        lambda x: x - timedelta(days=x.weekday())  # 关键修改2：直接按周一分组
    )
    
    print("数据分类中")

    bond_order = ['国债-新债', '国债-老债', '政策性金融债-新债', '政策性金融债-老债',
                  '中期票据', '短期/超短期融资券', '企业债', '地方政府债',
                  '同业存单', '资产支持证券']
    df['债券类型'] = pd.Categorical(
        df['债券类型'].fillna('二永债'),
        categories=bond_order,
        ordered=True
    )
    if jigou_mapping:
        df['机构类型'] = df['机构类型'].map(jigou_mapping).fillna('其他')
        bond_order = ['大行', '股份行', '城商行','证券', '保险','农商行','基金','理财','外资','贷基','其他']
        df['机构类型'] = pd.Categorical(
            df['机构类型'].fillna('其他'),
            categories=bond_order,
            ordered=True
        )
    print("开始创建新表格")

    # 创建透视表
    pivot_df = pd.pivot_table(
        df,
        values='净买入交易量（亿元）',
        index=['债券类型','周分组','期限'],
        columns='机构类型',
        aggfunc='sum',
        fill_value=0,
        observed=False#注意下，改过了
    )
    
    # 格式化周分组显示
    pivot_df.index = pivot_df.index.set_levels([
        pivot_df.index.levels[0],
        pivot_df.index.levels[1],
        pivot_df.index.levels[2]
    ])
    
    pivot_df.to_excel(output_file)
    print('新周表生成完成')
    
def create_colored_databar_graph_time(input_file,inst_name="",output_dir=""):
    plt.rcParams['legend.loc'] = 'upper right'  # 默认图例位置
    plt.rcParams['legend.frameon'] = False  # 优化渲染性能
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 微软雅黑
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    print("提取数据中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    df = df.ffill()
    
    print("机构数据筛选")
    if '机构类型' in df.columns and inst_name:
        df = df[df['机构类型'] == inst_name]
    
    if inst_name:
        if inst_name in df.columns:
            # 筛选关键列
            df = df[['债券类型', '周分组', '期限', inst_name]]
            # 转换目标列数据并重命名
            df['净买入交易量（亿元）'] = safe_convert_to_numeric(df[inst_name])
            df = df.drop(columns=[inst_name])  # 移除原列
        else:
            print(f"警告：未找到机构'{inst_name}'")
            return -1

    # 按周分组聚合
    weekly_df = df.groupby('周分组').agg({
        '净买入交易量（亿元）': 'sum',
        '期限': lambda x: x.mode()[0] if not x.empty else pd.NA
    }).reset_index()
    
    term_colors = {
    '≦1Y': "#FF0000", 
    '1-3Y': "#FF9100",
    '3-5Y': "#B8AF0E",
    '5-7Y': "#18B418",
    '7-10Y': "#356A66",
    '10-15Y': "#001EFF",
    '15-20Y': "#431271",
    '20-30Y': "#D113CE",
    '>30Y': "#7D1474"
    }

    # 创建画布
    fig, ax = plt.subplots(figsize=(20, 8))
    
  # 处理每个交易日数据
    for date_val, date_str in zip(df['周分组'].unique(), df['周分组'].unique()):
        daily_data = df[df['周分组'] == date_str]
        pos_data = daily_data[daily_data['净买入交易量（亿元）'] >= 0]
        neg_data = daily_data[daily_data['净买入交易量（亿元）'] < 0]
        
        # 正值部分绘制（宽度设为1实现无间隙）
        bottom = 0
        for _, row in pos_data.sort_values('期限').iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'], 
                width=6,
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#545454"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']
        
        # 负值部分绘制
        bottom = 0
        for _, row in neg_data.sort_values('期限', ascending=False).iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'],
                width=6, 
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#4D4C4C"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']


    # 图表基础美化配置
    ax.axhline(0, color='black', linewidth=0.8, linestyle='--')  # 虚线零线
    ax.set_ylabel('净买入值（亿元）', fontsize=12, fontweight='bold', color='#333333')
    ax.set_xlabel('日期', fontsize=12, fontweight='bold', color='#333333')
    

    title_name=inst_name+'债券期限结构交易分析'
    # 标题和坐标轴优化
    plt.title(title_name, 
            pad=20, 
            fontsize=14,
            fontweight='bold')
    plt.xticks(rotation=45, ha='right')  # 标签右对齐
    plt.grid(axis='y', linestyle='--', alpha=0.5)  # 横向网格线


    # 修正图例部分（关键修改点）
    handles = [plt.Rectangle((0,0),1,1, color=color, label=term) 
              for term, color in term_colors.items()]
    ax.legend(handles=handles, 
             title='期限分类',
             bbox_to_anchor=(1.02, 1),
             loc='upper left',
             borderaxespad=0.,
                frameon=False)
    
    #考虑移除
    ax.tick_params(axis='x', labelbottom=False)  # :ml-citation{ref="6,11" data="citationList"}
    
    # 提取'周分组'列的左侧日期（格式：2023/01/02-2023/01/06）
    date_vals = pd.to_datetime(pos_data['周分组'], format='%Y/%m/%d')

    # 计算动态刻度间隔（确保至少3个月，不超过数据范围）
    min_date, max_date = np.min(date_vals), np.max(date_vals)
    month_span = (max_date.year - min_date.year)*12 + (max_date.month - min_date.month)
    interval = max(3, month_span // 8)  # 自动调整间隔，保证最多显示8个标签

    # 配置刻度
    ax.xaxis.set_major_locator(MonthLocator(interval=interval))
    ax.xaxis.set_major_formatter(DateFormatter("%Y-%m"))
    ax.tick_params(axis='x', which='major', labelbottom=True, rotation=45)  # 

    plt.tight_layout(rect=[0.08, 0, 0.85, 1])
    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)
    # 拼接完整路径
    full_path = os.path.join(output_dir, title_name)
    plt.savefig(full_path, dpi=200, bbox_inches='tight')
    plt.close()  # 关键：关闭图形避免内存泄漏
    
    print(title_name+"已保存")
    return title_name
    
def create_colored_databar_graph_time_jing(input_file,inst_name="",output_dir=""):
    plt.rcParams['legend.loc'] = 'upper right'  # 默认图例位置
    plt.rcParams['legend.frameon'] = False  # 优化渲染性能
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 微软雅黑
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    print("提取数据中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    df = df.ffill()
    
    print("机构数据筛选")
    if '机构类型' in df.columns and inst_name:
        df = df[df['机构类型'] == inst_name]
    
    if inst_name:
        if inst_name in df.columns:
            # 筛选关键列（移除债券类型）
            df = df[['周分组', '期限', inst_name]]
            # 转换并重命名列
            df['净买入交易量（亿元）'] = safe_convert_to_numeric(df[inst_name])
            df = df.drop(columns=[inst_name])
            
            # 按周分组和期限计算总和
            df = df.groupby(['周分组', '期限'])['净买入交易量（亿元）'].sum().reset_index()
        else:
            print(f"警告：未找到机构'{inst_name}'")
            return -1
    

    # 按周分组聚合
    weekly_df = df.groupby('周分组').agg({
        '净买入交易量（亿元）': 'sum',
        '期限': lambda x: x.mode()[0] if not x.empty else pd.NA
    }).reset_index()
    
    term_colors = {
    '≦1Y': "#FF0000", 
    '1-3Y': "#FF9100",
    '3-5Y': "#B8AF0E",
    '5-7Y': "#18B418",
    '7-10Y': "#356A66",
    '10-15Y': "#001EFF",
    '15-20Y': "#431271",
    '20-30Y': "#D113CE",
    '>30Y': "#7D1474"
    }

    # 创建画布
    fig, ax = plt.subplots(figsize=(20, 8))
    
  # 处理每个交易日数据
    for date_val, date_str in zip(df['周分组'].unique(), df['周分组'].unique()):
        daily_data = df[df['周分组'] == date_str]
        pos_data = daily_data[daily_data['净买入交易量（亿元）'] >= 0]
        neg_data = daily_data[daily_data['净买入交易量（亿元）'] < 0]
        
        # 正值部分绘制（宽度设为1实现无间隙）
        bottom = 0
        for _, row in pos_data.sort_values('期限').iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'], 
                width=6,
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#545454"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']
        
        # 负值部分绘制
        bottom = 0
        for _, row in neg_data.sort_values('期限', ascending=False).iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'],
                width=6, 
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#4D4C4C"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']


    # 图表基础美化配置
    ax.axhline(0, color='black', linewidth=0.8, linestyle='--')  # 虚线零线
    ax.set_ylabel('净买入值（亿元）', fontsize=12, fontweight='bold', color='#333333')
    ax.set_xlabel('日期', fontsize=12, fontweight='bold', color='#333333')
    

    title_name=inst_name+'债券期限结构交易分析(净)'
    # 标题和坐标轴优化
    plt.title(title_name, 
            pad=20, 
            fontsize=14,
            fontweight='bold')
    plt.xticks(rotation=45, ha='right')  # 标签右对齐
    plt.grid(axis='y', linestyle='--', alpha=0.5)  # 横向网格线


    # 修正图例部分（关键修改点）
    handles = [plt.Rectangle((0,0),1,1, color=color, label=term) 
              for term, color in term_colors.items()]
    ax.legend(handles=handles, 
             title='期限分类',
             bbox_to_anchor=(1.02, 1),
             loc='upper left',
             borderaxespad=0.,
                frameon=False)
    
    #考虑移除
    ax.tick_params(axis='x', labelbottom=False)  # :ml-citation{ref="6,11" data="citationList"}
    
    # 提取'周分组'列的左侧日期（格式：2023/01/02-2023/01/06）
    date_vals = pd.to_datetime(pos_data['周分组'], format='%Y/%m/%d')

    # 计算动态刻度间隔（确保至少3个月，不超过数据范围）
    min_date, max_date = np.min(date_vals), np.max(date_vals)
    month_span = (max_date.year - min_date.year)*12 + (max_date.month - min_date.month)
    interval = max(3, month_span // 8)  # 自动调整间隔，保证最多显示8个标签

    # 配置刻度
    ax.xaxis.set_major_locator(MonthLocator(interval=interval))
    ax.xaxis.set_major_formatter(DateFormatter("%Y-%m"))
    ax.tick_params(axis='x', which='major', labelbottom=True, rotation=45)  # 

    plt.tight_layout(rect=[0.08, 0, 0.85, 1])
    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)
    # 拼接完整路径
    full_path = os.path.join(output_dir, title_name)
    plt.savefig(full_path, dpi=200, bbox_inches='tight')
    plt.close()  # 关键：关闭图形避免内存泄漏
    
    print(title_name+"已保存")
    return title_name
    
def create_colored_databar_graph_lilv(input_file,inst_name="",zhaiquan_mapping=None,output_dir=""):
    plt.rcParams['legend.loc'] = 'upper right'  # 默认图例位置
    plt.rcParams['legend.frameon'] = False  # 优化渲染性能
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 微软雅黑
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    print("提取数据中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    df = df.ffill()
    
    if zhaiquan_mapping:
        # 先创建包含所有利率债类型的列表
        rate_bond_types = [k for k,v in zhaiquan_mapping.items() if v == '利率债']

        # 使用isin筛选
        df = df[df['债券类型'].isin(rate_bond_types)]

    
    print("机构数据筛选")
    if '机构类型' in df.columns and inst_name:
        df = df[df['机构类型'] == inst_name]
    
    if inst_name:
        if inst_name in df.columns:
            # 筛选关键列
            df = df[['债券类型', '周分组', '期限', inst_name]]
            # 转换目标列数据并重命名
            df['净买入交易量（亿元）'] = safe_convert_to_numeric(df[inst_name])
            df = df.drop(columns=[inst_name])  # 移除原列
        else:
            print(f"警告：未找到机构'{inst_name}'")
            return -1

    # 按周分组聚合
    weekly_df = df.groupby('周分组').agg({
        '净买入交易量（亿元）': 'sum',
        '期限': lambda x: x.mode()[0] if not x.empty else pd.NA
    }).reset_index()
    
    term_colors = {
    '≦1Y': "#FF0000", 
    '1-3Y': "#FF9100",
    '3-5Y': "#B8AF0E",
    '5-7Y': "#18B418",
    '7-10Y': "#356A66",
    '10-15Y': "#001EFF",
    '15-20Y': "#431271",
    '20-30Y': "#D113CE",
    '>30Y': "#7D1474"
    }

    # 创建画布
    fig, ax = plt.subplots(figsize=(20, 8))
    
  # 处理每个交易日数据
    for date_val, date_str in zip(df['周分组'].unique(), df['周分组'].unique()):
        daily_data = df[df['周分组'] == date_str]
        pos_data = daily_data[daily_data['净买入交易量（亿元）'] >= 0]
        neg_data = daily_data[daily_data['净买入交易量（亿元）'] < 0]
        
        # 正值部分绘制（宽度设为1实现无间隙）
        bottom = 0
        for _, row in pos_data.sort_values('期限').iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'], 
                width=6,
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#545454"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']
        
        # 负值部分绘制
        bottom = 0
        for _, row in neg_data.sort_values('期限', ascending=False).iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'],
                width=6, 
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#4D4C4C"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']


    # 图表基础美化配置
    ax.axhline(0, color='black', linewidth=0.8, linestyle='--')  # 虚线零线
    ax.set_ylabel('净买入值（亿元）', fontsize=12, fontweight='bold', color='#333333')
    ax.set_xlabel('日期', fontsize=12, fontweight='bold', color='#333333')
    

    title_name=inst_name+'债券期限结构交易分析（利率债）'
    # 标题和坐标轴优化
    plt.title(title_name, 
            pad=20, 
            fontsize=14,
            fontweight='bold')
    plt.xticks(rotation=45, ha='right')  # 标签右对齐
    plt.grid(axis='y', linestyle='--', alpha=0.5)  # 横向网格线


    # 修正图例部分（关键修改点）
    handles = [plt.Rectangle((0,0),1,1, color=color, label=term) 
              for term, color in term_colors.items()]
    ax.legend(handles=handles, 
             title='期限分类',
             bbox_to_anchor=(1.02, 1),
             loc='upper left',
             borderaxespad=0.,
                frameon=False)
    
    #考虑移除
    ax.tick_params(axis='x', labelbottom=False)  # :ml-citation{ref="6,11" data="citationList"}
    
    # 提取'周分组'列的左侧日期（格式：2023/01/02-2023/01/06）
    date_vals = pd.to_datetime(pos_data['周分组'], format='%Y/%m/%d')

    # 计算动态刻度间隔（确保至少3个月，不超过数据范围）
    min_date, max_date = np.min(date_vals), np.max(date_vals)
    month_span = (max_date.year - min_date.year)*12 + (max_date.month - min_date.month)
    interval = max(3, month_span // 8)  # 自动调整间隔，保证最多显示8个标签

    # 配置刻度
    ax.xaxis.set_major_locator(MonthLocator(interval=interval))
    ax.xaxis.set_major_formatter(DateFormatter("%Y-%m"))
    ax.tick_params(axis='x', which='major', labelbottom=True, rotation=45)  # 

    plt.tight_layout(rect=[0.08, 0, 0.85, 1])
    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)
    # 拼接完整路径
    full_path = os.path.join(output_dir, title_name)
    plt.savefig(full_path, dpi=200, bbox_inches='tight')
    plt.close()  # 关键：关闭图形避免内存泄漏
    
    print(title_name+"已保存")
    return title_name
    

def create_colored_databar_graph_jinglilv(input_file,inst_name="",zhaiquan_mapping=None,output_dir=""):
    plt.rcParams['legend.loc'] = 'upper right'  # 默认图例位置
    plt.rcParams['legend.frameon'] = False  # 优化渲染性能
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 微软雅黑
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    print("提取数据中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    df = df.ffill()
    
    if zhaiquan_mapping:
        # 先创建包含所有利率债类型的列表
        rate_bond_types = [k for k,v in zhaiquan_mapping.items() if v == '利率债']

        # 使用isin筛选
        df = df[df['债券类型'].isin(rate_bond_types)]

    
    print("机构数据筛选")
    if '机构类型' in df.columns and inst_name:
        df = df[df['机构类型'] == inst_name]
    
    if inst_name:
        if inst_name in df.columns:
            # 筛选关键列（移除债券类型）
            df = df[['周分组', '期限', inst_name]]
            # 转换并重命名列
            df['净买入交易量（亿元）'] = safe_convert_to_numeric(df[inst_name])
            df = df.drop(columns=[inst_name])
            
            # 按周分组和期限计算总和
            df = df.groupby(['周分组', '期限'])['净买入交易量（亿元）'].sum().reset_index()
        else:
            print(f"警告：未找到机构'{inst_name}'")
            return -1
    
    term_colors = {
    '≦1Y': "#FF0000", 
    '1-3Y': "#FF9100",
    '3-5Y': "#B8AF0E",
    '5-7Y': "#18B418",
    '7-10Y': "#356A66",
    '10-15Y': "#001EFF",
    '15-20Y': "#431271",
    '20-30Y': "#D113CE",
    '>30Y': "#7D1474"
    }

    
    # 创建画布
    fig, ax = plt.subplots(figsize=(20, 8))
    
    
  # 处理每个交易日数据
    for date_val, date_str in zip(df['周分组'].unique(), df['周分组'].unique()):
        daily_data = df[df['周分组'] == date_str]
        pos_data = daily_data[daily_data['净买入交易量（亿元）'] >= 0]
        neg_data = daily_data[daily_data['净买入交易量（亿元）'] < 0]
        
        # 正值部分绘制（宽度设为1实现无间隙）
        bottom = 0
        for _, row in pos_data.sort_values('期限').iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'], 
                width=6,
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#545454"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']
        
        # 负值部分绘制
        bottom = 0
        for _, row in neg_data.sort_values('期限', ascending=False).iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'],
                width=6, 
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#4D4C4C"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']


    # 图表基础美化配置
    ax.axhline(0, color='black', linewidth=0.8, linestyle='--')  # 虚线零线
    ax.set_ylabel('净买入值（亿元）', fontsize=12, fontweight='bold', color='#333333')
    ax.set_xlabel('日期', fontsize=12, fontweight='bold', color='#333333')
    

    title_name=inst_name+'债券期限结构交易分析（利率债）(净)'
    # 标题和坐标轴优化
    plt.title(title_name, 
            pad=20, 
            fontsize=14,
            fontweight='bold')
    plt.xticks(rotation=45, ha='right')  # 标签右对齐
    plt.grid(axis='y', linestyle='--', alpha=0.5)  # 横向网格线


    # 修正图例部分（关键修改点）
    handles = [plt.Rectangle((0,0),1,1, color=color, label=term) 
              for term, color in term_colors.items()]
    ax.legend(handles=handles, 
             title='期限分类',
             bbox_to_anchor=(1.02, 1),
             loc='upper left',
             borderaxespad=0.,
                frameon=False)
    
    #考虑移除
    ax.tick_params(axis='x', labelbottom=False)  # :ml-citation{ref="6,11" data="citationList"}
    
    # 提取'周分组'列的左侧日期（格式：2023/01/02-2023/01/06）
    date_vals = pd.to_datetime(pos_data['周分组'], format='%Y/%m/%d')

    # 计算动态刻度间隔（确保至少3个月，不超过数据范围）
    min_date, max_date = np.min(date_vals), np.max(date_vals)
    month_span = (max_date.year - min_date.year)*12 + (max_date.month - min_date.month)
    interval = max(3, month_span // 8)  # 自动调整间隔，保证最多显示8个标签

    # 配置刻度
    ax.xaxis.set_major_locator(MonthLocator(interval=interval))
    ax.xaxis.set_major_formatter(DateFormatter("%Y-%m"))
    ax.tick_params(axis='x', which='major', labelbottom=True, rotation=45)  # 

    plt.tight_layout(rect=[0.08, 0, 0.85, 1])
    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)
    # 拼接完整路径
    full_path = os.path.join(output_dir, title_name)
    plt.savefig(full_path, dpi=200, bbox_inches='tight')
    plt.close()  # 关键：关闭图形避免内存泄漏
    
    print(title_name+"已保存")
    return title_name

def create_colored_databar_graph_single_bond(input_file,inst_name="",bond_name="",xinjiu_mapping=None,output_dir=""):
    plt.rcParams['legend.loc'] = 'upper right'  # 默认图例位置
    plt.rcParams['legend.frameon'] = False  # 优化渲染性能
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 微软雅黑
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    print("提取数据中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    df = df.ffill()
    
    if zhaiquan_mapping:
        # 先创建包含所有利率债类型的列表
        rate_bond_types = [k for k,v in xinjiu_mapping.items() if v == bond_name]

        # 使用isin筛选
        df = df[df['债券类型'].isin(rate_bond_types)]

    
    print("机构数据筛选")
    if '机构类型' in df.columns and inst_name:
        df = df[df['机构类型'] == inst_name]
    
    if inst_name:
        if inst_name in df.columns:
            # 筛选关键列（移除债券类型）
            df = df[['周分组', '期限', inst_name]]
            # 转换并重命名列
            df['净买入交易量（亿元）'] = safe_convert_to_numeric(df[inst_name])
            df = df.drop(columns=[inst_name])
            
            # 按周分组和期限计算总和
            df = df.groupby(['周分组', '期限'])['净买入交易量（亿元）'].sum().reset_index()
        else:
            print(f"警告：未找到机构'{inst_name}'")
            return -1
    
    term_colors = {
    '≦1Y': "#FF0000", 
    '1-3Y': "#FF9100",
    '3-5Y': "#B8AF0E",
    '5-7Y': "#18B418",
    '7-10Y': "#356A66",
    '10-15Y': "#001EFF",
    '15-20Y': "#431271",
    '20-30Y': "#D113CE",
    '>30Y': "#7D1474"
    }
    
    # 创建画布
    fig, ax = plt.subplots(figsize=(20, 8))
    
    
  # 处理每个交易日数据
    for date_val, date_str in zip(df['周分组'].unique(), df['周分组'].unique()):
        daily_data = df[df['周分组'] == date_str]
        pos_data = daily_data[daily_data['净买入交易量（亿元）'] >= 0]
        neg_data = daily_data[daily_data['净买入交易量（亿元）'] < 0]
        
        # 正值部分绘制（宽度设为1实现无间隙）
        bottom = 0
        for _, row in pos_data.sort_values('期限').iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'], 
                width=6,
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#545454"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']
        
        # 负值部分绘制
        bottom = 0
        for _, row in neg_data.sort_values('期限', ascending=False).iterrows():
            ax.bar(date_val, row['净买入交易量（亿元）'],
                width=6, 
                bottom=bottom, 
                color=term_colors.get(row['期限'], "#4D4C4C"),
                edgecolor='none')  # 关键修改点
            bottom += row['净买入交易量（亿元）']


    # 图表基础美化配置
    ax.axhline(0, color='black', linewidth=0.8, linestyle='--')  # 虚线零线
    ax.set_ylabel('净买入值（亿元）', fontsize=12, fontweight='bold', color='#333333')
    ax.set_xlabel('日期', fontsize=12, fontweight='bold', color='#333333')
    

    title_name=inst_name+'债券期限结构交易分析（'+bond_name+'）'
    # 标题和坐标轴优化
    plt.title(title_name, 
            pad=20, 
            fontsize=14,
            fontweight='bold')
    plt.xticks(rotation=45, ha='right')  # 标签右对齐
    plt.grid(axis='y', linestyle='--', alpha=0.5)  # 横向网格线


    # 修正图例部分（关键修改点）
    handles = [plt.Rectangle((0,0),1,1, color=color, label=term) 
              for term, color in term_colors.items()]
    ax.legend(handles=handles, 
             title='期限分类',
             bbox_to_anchor=(1.02, 1),
             loc='upper left',
             borderaxespad=0.,
                frameon=False)
    
    #考虑移除
    ax.tick_params(axis='x', labelbottom=False)  # :ml-citation{ref="6,11" data="citationList"}
    
    # 提取'周分组'列的左侧日期（格式：2023/01/02-2023/01/06）
    date_vals = pd.to_datetime(pos_data['周分组'], format='%Y/%m/%d')

    # 计算动态刻度间隔（确保至少3个月，不超过数据范围）
    min_date, max_date = np.min(date_vals), np.max(date_vals)
    month_span = (max_date.year - min_date.year)*12 + (max_date.month - min_date.month)
    interval = max(3, month_span // 8)  # 自动调整间隔，保证最多显示8个标签

    # 配置刻度
    ax.xaxis.set_major_locator(MonthLocator(interval=interval))
    ax.xaxis.set_major_formatter(DateFormatter("%Y-%m"))
    ax.tick_params(axis='x', which='major', labelbottom=True, rotation=45)  # 

    plt.tight_layout(rect=[0.08, 0, 0.85, 1])
    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)
    # 拼接完整路径
    full_path = os.path.join(output_dir, title_name)
    plt.savefig(full_path, dpi=200, bbox_inches='tight')
    plt.close()  # 关键：关闭图形避免内存泄漏
    
    print(title_name+"已保存")
    return title_name
    
def apply_grouped_data_bars(file_path, group_size=128):
    print('分组数据着色开始')
    """
    使用xlwings实现分组数据条着色
    参数:
        file_path: Excel文件路径
        group_size: 每组行数(默认128)
    """
    #打开Excel
    app = xw.App(visible=False)
    app.screen_updating = False
    app.calculation = 'manual'  # 禁止公式重算
    app.display_alerts = False  # 关闭提示框
    wb = xw.Book(file_path)
    ws = wb.sheets.active
    full_range = f"{'C'}{2}:{'M'}{600}"
    ws.range(full_range).api.FormatConditions.Delete()

    # 动态获取最后一行
    last_row = ws.range('C' + str(ws.cells.last_cell.row)).end('up').row
    print(f"检测到数据总行数: {last_row}")
    
    # 处理C到M列(3-13列)
    for col in range(3, 14):
        col_letter = chr(64 + col)  # 转换为字母
        print (f"已完成列{col_letter}")
        # 按组处理(2-129,130-257,...)
        for start_row in range(2, last_row + 1, group_size):
            end_row = min(start_row + group_size - 1, last_row)
            cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            
            # 添加数据条
            data_bar = ws.range(cell_range).api.FormatConditions.AddDatabar()
            
            # 设置数据条格式
            data_bar.BarColor.Color = 0x0000FF  # 红色（虽然看起来是蓝色，但是显示出来是红色？我也不知道为什么）
            data_bar.BarFillType = 0  # xlDataBarFillGradient
            data_bar.BarBorder.Type = 0  # xlDataBarBorderSolid
            
            # 设置轴格式
            data_bar.AxisPosition = 0  # xlDataBarAxisAutomatic
            data_bar.AxisColor.Color = 0x000000  # 黑色
            
            # 设置负数格式
            neg_format = data_bar.NegativeBarFormat
            neg_format.ColorType = 0  # xlDataBarColor
            neg_format.Color.Color = 0x00FF00  # 绿色
            neg_format.BorderColorType = 0  # xlDataBarColor
    #保存并关闭
    wb.save()
    wb.close()
    app.quit()
    print(f"已完成按行分组的C-M列数据条设置，结果已保存到: {file_path}")

    
def safe_convert_to_numeric(series):
    numeric_series = pd.to_numeric(series, errors='coerce')
    if numeric_series.isna().any():
        cleaned = series.str.replace(',', '').str.extract(r'([-+]?\d*\.?\d+)')[0]
        numeric_series = pd.to_numeric(cleaned, errors='coerce')
    return numeric_series.fillna(0)

def get_week_range(date_obj):
    monday = date_obj - timedelta(days=date_obj.weekday())
    friday = monday + timedelta(days=4)
    return f"{monday.year}/{monday.month:02d}/{monday.day:02d}-{friday.year}/{friday.month:02d}/{friday.day:02d}"

def process_data(input_file, output_file, zhaiquan_mapping=None, jigou_mapping=None):
    # 数据读取与预处理
    print("数据提取中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)
    
    if '净买入交易量（亿元）' in df.columns:
        df['净买入交易量（亿元）'] = safe_convert_to_numeric(df['净买入交易量（亿元）'])
    
    df['交易日期'] = pd.to_datetime(df['交易日期'], format='%Y/%m/%d', errors='coerce')
    df = df[df['交易日期'].dt.weekday < 5]  # 关键修改1：过滤周末
    
    df['周分组'] = df['交易日期'].apply(
        lambda x: x - timedelta(days=x.weekday())  # 关键修改2：直接按周一分组
    )
    
    print("数据分类中")

    # 应用映射规则
    if zhaiquan_mapping:
        df['债券类型'] = df['债券类型'].map(zhaiquan_mapping).fillna('二永债')
        bond_order = ['利率债', '信用债', '同业存单', '二永债']
        df['债券类型'] = pd.Categorical(
            df['债券类型'].fillna('二永债'),
            categories=bond_order,
            ordered=True
        )
    if jigou_mapping:
        df['机构类型'] = df['机构类型'].map(jigou_mapping).fillna('其他')
        bond_order = ['大行', '股份行', '城商行','证券', '保险','农商行','基金','理财','外资','贷基','其他']
        df['机构类型'] = pd.Categorical(
            df['机构类型'].fillna('其他'),
            categories=bond_order,
            ordered=True
        )
    print("开始创建新表格")

    # 创建透视表
    pivot_df = pd.pivot_table(
        df,
        values='净买入交易量（亿元）',
        index=['债券类型','周分组'],
        columns='机构类型',
        aggfunc='sum',
        fill_value=0,
        observed=False#注意下，改过了
    )
    
    # 格式化周分组显示
    pivot_df.index = pivot_df.index.set_levels([
        pivot_df.index.levels[0],
        pivot_df.index.levels[1].map(lambda x: get_week_range(x))
    ])
    
    pivot_df.to_excel(output_file)
    print('新表生成完成')




def filter_to_excel_long_term(input_file, output_file,zhaiquan_mapping):
    print("提取数据中")
    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()
    df.replace('--', pd.NA, inplace=True)

    
    print("期限筛选")
    long_term = ['10-15Y', '15-20Y', '20-30Y', '>30Y']
    df = df[df['期限'].isin(long_term)]

    if zhaiquan_mapping:
        rate_bond_types = [k for k,v in zhaiquan_mapping.items() if v == '利率债']
        df = df[df['债券类型'].isin(rate_bond_types)]
        
        institutions = ['大型商业银行/政策性银行','保险公司','基金公司及产品','农村金融机构']
        for inst in institutions:
            df[f'{inst}_利率债'] = np.where(
                df['机构类型']==inst, 
                df['净买入交易量（亿元）'],
                0
            )
            
        local_bond_types = ['地方政府债']
        mask = (df['债券类型'].isin(local_bond_types)) & (df['机构类型']=='保险公司')
        df['保险地方债'] = np.where(mask, df['净买入交易量（亿元）'], 0)
        
    # 修正后的列删除逻辑
    cols_to_drop = ['机构类型','净买入交易量（亿元）','买入交易量（亿元）','卖出交易量（亿元）','期限']
    existing_cols = [col for col in cols_to_drop if col in df.columns]
    if existing_cols:
        df = df.drop(columns=existing_cols)
    else:
        print("警告: 未找到任何待删除列")
    
    cols = ['大型商业银行/政策性银行_利率债','保险公司_利率债','基金公司及产品_利率债','农村金融机构_利率债','保险地方债']
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce')
    df[cols] = df[cols].fillna(0)
    
    result = df.groupby('交易日期')[cols].sum().reset_index()
    
    result['交易日期'] = pd.to_datetime(result['交易日期']).dt.date

    result.to_excel(output_file, index=False)
    print(f"处理完成，结果已保存至 {output_file}")


  

'''从这里往下可以改，上面最好不要改'''
if __name__ == '__main__':
    
    #债券分类，'债券：种类'
    zhaiquan_mapping = {'国债-新债':'利率债','地方政府债':'利率债'
                ,'国债-老债':'利率债','政策性金融债-新债':'利率债'
                ,'政策性金融债-老债':'利率债', '中期票据':'信用债'
                , '短期/超短期融资券':'信用债', '企业债':'信用债'
                , '资产支持证券':'信用债','同业存单':'同业存单'}
    #机构，'机构：种类'
    jigou_mapping = {
        '保险公司':'保险', '城市商业银行':'城商行',
        '大型商业银行/政策性银行':'大行', '股份制商业银行':'股份行',
        '货币市场基金':'贷基', '外资银行':'外资',
        '基金公司及产品':'基金', '理财子公司及理财类产品':'理财',
        '农村金融机构':'农商行', '证券公司':'证券'
    }
    #取消新老债，例如新老国债全归为国债
    xinjiu_mapping = {'国债-新债':'国债','地方政府债':'地方政府债'
            ,'国债-老债':'国债','政策性金融债-新债':'政策性金融债'
            ,'政策性金融债-老债':'政策性金融债', '中期票据':'中期票据'
            , '短期/超短期融资券':'短期/超短期融资券', '企业债':'企业债'
            , '资产支持证券':'资产支持证券','同业存单':'同业存单'
        
    }
    
    '''可更改，文件路径'''
    raw_input_folder='C:/Users/Linda/合并测试/日数据'  # 替换为您希望合并的文件的所属文件夹路径
    raw_output_file='C:/Users/Linda/合并测试/合并数据/2023-25日数据v3.xlsx'  # 替换为输出文件夹路径/文件名称.xlsx,请确保文件夹里没有相应的文件名称
    input_file=raw_output_file#等于处理过的周数据
    output_file='C:/Users/Linda/合并测试/合并数据/2023-25季度总结v14.xlsx'
    output_file_week='C:/Users/Linda/合并测试/合并数据/2023-25周数据v5.xlsx'
    output_file_long_term='C:/Users/Linda/合并测试/合并数据/十年以上利率债v5.xlsx'
    
    '''单做表的债券种类，与机构种类'''
    important_bond=['国债','地方政府债','政策性金融债']
    important_inst=['保险','基金','农商行','大行','外资']
    
    '''图片保存文件夹与excel表格地点，请确保文件夹与excel为空'''
    image_save='C:/Users/Linda/合并测试/图片'#请确保文件夹为空
    excel_path='C:/Users/Linda/合并测试/图片/总览.xlsx'
    
    #excel表格中的图片大小，可更改
    img_width=700
    img_height=400
    
    #请确保所有斜杠都是这样的/
        #否则可能报错
    '''如果数据更新，从这里开始跑'''
    merge_second_sheets(raw_input_folder, raw_output_file)
    
    '''如果input_file已更新，可以从这里开始跑(合并完成的N年数据)'''
    #excel表格着色
    process_data(input_file, output_file, zhaiquan_mapping, jigou_mapping)
    apply_grouped_data_bars(output_file)
    daily_to_weekly(input_file, output_file_week)
    
    
    '''如果output_file_week已更新，可以从这里开始跑'''
    
    i=0
    column=['A','B','C','D','E','F','G','H','I','J','K','L','M']
    wb = Workbook()  # 创建新工作簿
    wb.save(excel_path)  # 保存为xlsx格式
    set_excel_dimensions(excel_path, 50, img_height, column, img_width)
    print("0/"+str(len(important_inst)))
    
    #画表
    for inst in important_inst:

        filename=create_colored_databar_graph_time(output_file_week,inst,image_save)
        filename=create_colored_databar_graph_time_jing(output_file_week,inst,image_save)
        filename=create_colored_databar_graph_lilv(output_file_week,inst,zhaiquan_mapping,image_save)
        filename=create_colored_databar_graph_jinglilv(output_file_week,inst,zhaiquan_mapping,image_save)

        
        for bond in important_bond:
            filename=create_colored_databar_graph_single_bond(output_file_week,inst,bond,xinjiu_mapping,image_save)
        i+=1
        print(str(i)+"/"+str(len(important_inst)))
        
    print("完成")
    
    
    '''如果文件夹地点已生成数据，可以从这里跑'''
    insert_images_to_excel(excel_path, image_save, "A1", img_width, img_height,4+len(important_bond))
    
    
    '''10年以上利率债数据，excel生成'''
    filter_to_excel_long_term(raw_output_file,output_file_long_term,zhaiquan_mapping)
    
    
    #ctrl + /   注释
    #跑程序的时候要确保相关文件没有在excel里面开着